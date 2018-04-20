from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re
import pprint
import io
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage


def convert_pdf_to_txt(path):
    rsrcmgr = PDFResourceManager()
    retstr = io.StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                  password=password,
                                  caching=caching,
                                  check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    return text

file = open("out.txt","w")
result = convert_pdf_to_txt('newData.pdf')
file.write(str(result))
file.close()

wb = load_workbook('ourData.xlsx')
sheet = wb['Sheet1']

cell_range=sheet['J3':sheet.max_row]


for each_column in cell_range:
    for cell in each_column:
        searchCode = str(cell.value)
        if searchCode in open('out.txt').read():
            myNum = column_index_from_string(cell.column)
            newColCart = myNum - 6
            newColName = myNum - 3
            outputfile = open('newFile.txt','a')
            name = sheet.cell(cell.row, newColName).value
            cart = sheet.cell(cell.row, newColCart).value
            mydict = {'name': name, 'cart': cart, 'ID': searchCode}
            pprint.pformat(mydict)
            outputfile.write(pprint.pformat(mydict) + '\n')
            
outputfile.close()
